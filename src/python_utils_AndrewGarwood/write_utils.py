from typing import List, Dict, Any, Literal
import os
from dotenv import load_dotenv
from datetime import datetime
import pandas as pd
from pandas import DataFrame
from typing import Dict, List, Literal
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from .file_utils import validate_file_extension

load_dotenv()
DEFAULT_LOG: str = os.getenv('DEFAULT_LOG_FILE')
FIELD_UPDATE_LOG: str = os.getenv('FIELD_UPDATE_LOG')

# like a worse version of json pretty printing...
def print_group(
    label: str = f'Print Group {datetime.now()}', 
    data: List[str] = [], 
    indent: int = 0,
    print_to_console: bool = True, 
    log_path: str = None,
    mode: Literal['w', 'a'] = 'a'
) -> None:
    timestamp: str = f"{datetime.now()}".split('.')[0]
    if not log_path and print_to_console:
        print("\t"*indent + f"({timestamp}) {label}")
        for item in data:
            print("\t"*(indent+1) + f"{item}")
    elif log_path and print_to_console:
        output_path = log_path if log_path and os.path.isfile(log_path) else DEFAULT_LOG
        with open(file=output_path, mode=mode) as f:
            f.write("\t"*indent + f"({timestamp}) {label}\n")
            for item in data:
                f.write("\t"*(indent+1) + f"{item}\n")
        print("\t"*indent + f"({timestamp}) {label}")
        for item in data:
            print("\t"*(indent+1) + f"{item}")
    else:
        output_path = log_path if log_path and os.path.isfile(log_path) else DEFAULT_LOG
        with open(file=output_path, mode=mode) as f:
            f.write("\t"*indent + f"({timestamp}) {label}\n")
            for item in data:
                f.write("\t"*(indent+1) + f"{item}\n")
                

def concatenate_dataframes_to_excel(
    dataframes: List[DataFrame], 
    orientation: Literal['horizontal', 'vertical', 'h', 'v', 'horiz', 'vert', 'row', 'col'],
    output_path: str, 
    sheet_name: str = 'Sheet1',
    empty_spaces: int = 1
) -> None:
    output_path += '.xlsx' if not output_path.endswith('.xlsx') else ''
    try:
        wb: Workbook = load_workbook(output_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet created with a new workbook

    ws: Worksheet = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = 'A2'  # Freeze the first row
    if orientation.lower() in ['horizontal', 'h', 'horiz', 'row']:
        current_col = 1
        for df in dataframes:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, current_col):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            current_col += df.shape[1] + empty_spaces
    elif orientation.lower() in ['vertical', 'v', 'vert', 'col']:
        current_row = 1
        for df in dataframes:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), current_row):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            current_row += df.shape[0] + empty_spaces + 1

    else:
        raise ValueError("Orientation must be either 'horizontal' or 'vertical'")
    wb.save(output_path)


def write_dataframes_to_excel(
    file_path: str, 
    df_dict: Dict[str, DataFrame]
) -> None:
    file_path += '.xlsx' if not file_path.endswith('.xlsx') else ''
    try: # Load the existing workbook  
        workbook: Workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames
    except FileNotFoundError: # If the file does not exist, create a new workbook
        workbook = None
        sheet_names = []
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df = pd.DataFrame({'': ['']})
            df.to_excel(writer, sheet_name='Sheet0', index=False, freeze_panes=(1, 0))
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        # writer.book = workbook
        for df_name, df in df_dict.items():
            # Ensure unique sheet names
            original_sheet_name = df_name
            index = 1
            while df_name in sheet_names:
                df_name = f"{original_sheet_name}{index}"
                index += 1
            df.to_excel(
                writer, sheet_name=df_name, freeze_panes=(1, 0),
                merge_cells=False, index=False
            )
            sheet_names.append(df_name)